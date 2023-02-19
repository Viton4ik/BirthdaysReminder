
import openpyxl

from datetime import datetime
import time
import calendar

from random import randint

import requests

# Telegram Token
TOKEN = 'YOUR_TOKEN' 
chat_id_public = [] # put list of your public chat_id here !!!
chat_id_private = [] # put list of your private chat_id here !!! 

year_months = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June', 7: 'July', 8: 'August',
                9: 'September', 10: 'October', 11: 'November', 12: 'December'}

today = datetime.today().date()
today_year = datetime.today().date().year
today_month = datetime.today().date().month
today_day = datetime.today().date().day

file = 'Birthdays.xlsx'

#1
def getBirthdaysList():
    # open file xlsx
    workbook = openpyxl.load_workbook(file, data_only=True) # data_only=True - to get a data from cell instead of formula
    worksheet = workbook.active
    
    personInfo = {}
    generalInfo = {}
    # first cell
    col = 10
    row = 1
    id = worksheet.cell(col, row).value
    while id is not None:
        personInfo["id"] = worksheet.cell(col, row).value
        personInfo["slava"] = worksheet.cell(col, row+2).value
        personInfo["stDay"] = worksheet.cell(col, row+3).value
        personInfo["birthday"] = worksheet.cell(col, row+4).value
        personInfo["phoneNumber"] = worksheet.cell(col, row+5).value
        personInfo["age"] = worksheet.cell(col, row+6).value
        personInfo["private"] = worksheet.cell(col, row+7).value
        personInfo["yearIsUnknown"] = worksheet.cell(col, row + 8).value
        generalInfo[worksheet.cell(col, row + 1).value] = personInfo
        personInfo = {}
        col += 1
        id = worksheet.cell(col, row).value
    return generalInfo

#2
def getBirthdayAlert():
    for person in getBirthdaysList():
        birthday_month = getBirthdaysList()[person]['birthday'].date().month
        birthday_day = getBirthdaysList()[person]['birthday'].date().day
        if today_month == birthday_month:
            if birthday_day - today_day == 1:
                if getBirthdaysList()[person]['private'] == 'true':
                    chat_id = chat_id_private
                else:
                    chat_id = chat_id_public
                ageInfo = f"By the way, that young person will turn '<b>{int(getBirthdaysList()[person]['age'])+1}</b>' years tomorrow!" \
                    if getBirthdaysList()[person]['yearIsUnknown'] == 'false' else ''
                message = f"'<b>{person}</b>' is celebrating the birthday TOMORROW - {year_months[birthday_month]} of {birthday_day}! Don't forget it! \n{ageInfo}"
                for chat in chat_id:
                    url = f"https://api.telegram.org/bot{TOKEN}/sendMessage?chat_id={chat}&text={message}&parse_mode=HTML"
                    print(requests.get(url).json())
                    time.sleep(0.5)
            elif birthday_day - today_day == 0:
                if getBirthdaysList()[person]['private'] == 'true':
                    chat_id = chat_id_private
                else:
                    chat_id = chat_id_public
                ageInfo = f"By the way, that young person has turned '<b>{int(getBirthdaysList()[person]['age'])}</b>' years today!" \
                    if getBirthdaysList()[person]['yearIsUnknown'] == 'false' else ''
                message = f"'<b>{person}</b>' is celebrating the birthday TODAY - {year_months[birthday_month]} of {birthday_day}!\n{ageInfo}"
                for chat in chat_id:
                    file_ = 'vitya' if person == 'Victor Vetoshkin' else randint(1, 22)
                    file = {'photo': (f'Images/{file_}.jpg', open(f'Images/{file_}.jpg', 'rb'))}
                    data = {
                        'chat_id': chat,
                        'caption': message,
                        'parse_mode': 'HTML' # get telegram fonts
                    }
                    # response to API
                    response = requests.post(f"https://api.telegram.org/bot{TOKEN}/sendPhoto",
                                             data=data, files=file)
                    print(response.status_code)
                    time.sleep(0.5)
        elif birthday_day == 1 and today == calendar.monthrange(today_year, today_month)[1]:
            if getBirthdaysList()[person]['private'] == 'true':
                chat_id = chat_id_private
            else:
                chat_id = chat_id_public
            ageInfo = f"By the way, that young person will turn '<b>{int(getBirthdaysList()[person]['age'])}</b>' years tomorrow!" \
                if getBirthdaysList()[person]['yearIsUnknown'] == 'false' else ''
            message = f"'<b>{person}</b>' is celebrating the birthday TOMORROW - {year_months[birthday_month]} of {birthday_day}! Don't forget it! \n{ageInfo}"
            for chat in chat_id:
                url = f"https://api.telegram.org/bot{TOKEN}/sendMessage?chat_id={chat}&text={message}&parse_mode=HTML"
                print(requests.get(url).json())
                time.sleep(0.5)

#3
def getSlavaAlert():
    for person in getBirthdaysList():
        if getBirthdaysList()[person]['slava'] != 'None':
            slavaMonth = getBirthdaysList()[person]['slava'].date().month
            slavaDay = getBirthdaysList()[person]['slava'].date().day
            chat_id = chat_id_public
            if today_month == slavaMonth:
                if slavaDay - today_day == 1:
                    message = f"'<b>{person}</b>' is celebrating his slava day TOMORROW: {year_months[slavaMonth]} of {slavaDay} - <b>{getBirthdaysList()[person]['stDay']}</b>!"
                    for chat in chat_id:
                        url = f"https://api.telegram.org/bot{TOKEN}/sendMessage?chat_id={chat}&text={message}&parse_mode=HTML"
                        print(requests.get(url).json())
                        time.sleep(0.5)
                elif slavaDay - today_day == 0:
                    message = f"'<b>{person}</b>' is celebrating his slava day TODAY: {today} - <b>{getBirthdaysList()[person]['stDay']}</b>!"
                    for chat in chat_id:
                        url = f"https://api.telegram.org/bot{TOKEN}/sendMessage?chat_id={chat}&text={message}&parse_mode=HTML"
                        print(requests.get(url).json())
                        time.sleep(0.5)
                elif slavaDay - today_day == 4:
                    message = f"'<b>{person}</b>' is celebrating his slava day in 4 days: {year_months[slavaMonth]} of {slavaDay} - <b>{getBirthdaysList()[person]['stDay']}</b>!"
                    for chat in chat_id:
                        url = f"https://api.telegram.org/bot{TOKEN}/sendMessage?chat_id={chat}&text={message}&parse_mode=HTML"
                        print(requests.get(url).json())
                        time.sleep(0.5)

#4
def loop():
    timer_hours = 14
    timer_min = 6
    timer_sec = 00
    time_now = datetime.now().time()
    if time_now.hour == timer_hours and time_now.minute == timer_min and time_now.second == timer_sec:
        getBirthdayAlert()
        getSlavaAlert()

def main():
    while True:
        loop()

if __name__ == '__main__':
    main()
